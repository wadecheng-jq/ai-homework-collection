#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI作业收集系统 - 主服务程序
功能：网页表单提交、附件上传、Excel自动记录、截止时间锁定
作者：WorkBuddy for Wade
日期：2026-06-05
"""

import os
import sys
import io
import json
import datetime
from flask import Flask, request, render_template, jsonify, send_file, abort
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import hashlib

# ============ 配置区 ============
APP_DIR = os.path.dirname(os.path.abspath(__file__))
UPLOAD_DIR = os.path.join(APP_DIR, "static", "uploads")
EXCEL_PATH = os.path.join(APP_DIR, "submissions.xlsx")
TEMPLATE_DIR = os.path.normpath(os.path.join(APP_DIR, "templates"))

# 截止时间：2026-06-08 18:00:00
DEADLINE = datetime.datetime(2026, 6, 8, 18, 0, 0)

# 允许上传的文件扩展名
ALLOWED_EXTENSIONS = {
    'png', 'jpg', 'jpeg', 'gif', 'bmp', 'webp',  # 图片
    'pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx',  # Office
    'txt', 'md', 'csv', 'json', 'xml', 'html',  # 文本
    'zip', 'rar', '7z', 'tar', 'gz',  # 压缩包
    'mp4', 'avi', 'mov', 'wmv', 'flv',  # 视频
}

# ============ Flask 初始化 ============
app = Flask(__name__, template_folder=TEMPLATE_DIR, static_folder=os.path.join(APP_DIR, "static"))
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 最大50MB
app.config['UPLOAD_FOLDER'] = UPLOAD_DIR

# 确保目录存在
os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(TEMPLATE_DIR, exist_ok=True)


# ============ 工具函数 ============
def allowed_file(filename):
    """检查文件扩展名是否允许"""
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def get_submission_count():
    """获取当前提交人数"""
    if not os.path.exists(EXCEL_PATH):
        return 0
    try:
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
        return ws.max_row - 1 if ws.max_row > 1 else 0
    except Exception:
        return 0


def is_deadline_passed():
    """检查是否已过截止时间"""
    return datetime.datetime.now() > DEADLINE


def init_excel():
    """初始化Excel文件（如不存在则创建）"""
    if os.path.exists(EXCEL_PATH):
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "提交记录"

    # 表头样式
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name='微软雅黑', bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ["序号", "提交时间", "姓名", "圈子/部门", "AI工具/大模型", "作业文字内容", "附件文件名", "附件路径", "状态"]
    col_widths = [8, 20, 15, 20, 20, 60, 30, 50, 12]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 30
    wb.save(EXCEL_PATH)
    print(f"[INFO] 初始化Excel文件：{EXCEL_PATH}")


def append_submission(name, group, ai_tool, content, filename, filepath):
    """追加一条提交记录到Excel"""
    init_excel()
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active

    next_row = ws.max_row + 1
    now_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    data = [next_row - 1, now_str, name, group, ai_tool, content, filename, filepath, "已提交"]
    aligns = ["center", "center", "center", "center", "center", "left", "left", "left", "center"]
    wraps = [False, False, False, False, False, True, False, False, False]

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_idx, (value, align, wrap) in enumerate(zip(data, aligns, wraps), 1):
        cell = ws.cell(row=next_row, column=col_idx, value=value)
        cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
        cell.border = thin_border
        if next_row % 2 == 0:
            cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    ws.row_dimensions[next_row].height = 60 if wraps[5] else 25
    wb.save(EXCEL_PATH)
    return next_row - 1


def get_all_submissions():
    """获取所有提交记录"""
    if not os.path.exists(EXCEL_PATH):
        return []
    wb = load_workbook(EXCEL_PATH)
    ws = wb.active
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            records.append({
                "seq": row[0],
                "time": row[1],
                "name": row[2],
                "group": row[3],
                "ai_tool": row[4],
                "content": row[5],
                "filename": row[6],
                "filepath": row[7],
                "status": row[8],
            })
    return records


def get_submitted_names():
    """获取所有已提交人员的姓名集合（去重，忽略大小写和空格）"""
    records = get_all_submissions()
    names = set()
    for r in records:
        if r["name"]:
            names.add(r["name"].strip().lower())
    return names


# ============ 路由 ============

@app.route("/", methods=["GET"])
def index():
    """显示提交表单页面"""
    passed = is_deadline_passed()
    count = get_submission_count()
    return render_template(
        "form.html",
        deadline_passed=passed,
        deadline_str=DEADLINE.strftime("%Y年%m月%d日 %H:%M"),
        submission_count=count,
    )


@app.route("/submit", methods=["POST"])
def submit():
    """处理表单提交"""
    if is_deadline_passed():
        return jsonify({"success": False, "message": f"已超过截止时间（{DEADLINE.strftime('%Y-%m-%d %H:%M')}），提交已关闭。如有特殊情况请联系管理员。"}), 403

    try:
        name = request.form.get("name", "").strip()
        group = request.form.get("group", "").strip()
        content = request.form.get("content", "").strip()
        ai_tool = request.form.get("ai_tool", "").strip()

        if not name:
            return jsonify({"success": False, "message": "请输入姓名！"}), 400
        if not content:
            return jsonify({"success": False, "message": "请输入作业文字内容！"}), 400
        if not ai_tool:
            return jsonify({"success": False, "message": "请填写你使用的AI工具和大模型！"}), 400

        # 检查是否已提交（按姓名去重）
        submitted = get_submitted_names()
        if name.lower() in submitted:
            return jsonify({"success": False, "message": f"「{name}」已经提交过作业，无需重复提交。如需修改请联系管理员。"}), 400

        # 处理附件
        filename = ""
        filepath = ""
        if "file" in request.files:
            file = request.files["file"]
            if file and file.filename != "":
                if not allowed_file(file.filename):
                    return jsonify({"success": False, "message": "不支持的文件类型，请上传图片、Office文档、PDF或压缩包等常见格式。"}), 400

                # 生成安全文件名（保留原扩展名，用MD5避免重名）
                orig_name = secure_filename(file.filename)
                ext = orig_name.rsplit(".", 1)[-1] if "." in orig_name else ""
                name_hash = hashlib.md5(f"{name}{datetime.datetime.now()}".encode()).hexdigest()[:10]
                safe_name = f"{name_hash}.{ext}" if ext else f"{name_hash}.bin"

                save_path = os.path.join(UPLOAD_DIR, safe_name)
                file.save(save_path)
                filename = orig_name
                filepath = os.path.relpath(save_path, APP_DIR)

        # 写入Excel
        seq = append_submission(name, group, ai_tool, content, filename, filepath)
        print(f"[INFO] 收到提交 #{seq}：{name}（{group}）- AI工具：{ai_tool}")

        return jsonify({
            "success": True,
            "message": f"提交成功！感谢「{name}」的提交，序号 #{seq}。",
            "seq": seq,
        })

    except Exception as e:
        print(f"[ERROR] 提交处理失败：{e}")
        return jsonify({"success": False, "message": f"提交失败：{str(e)}"}), 500


@app.route("/api/status", methods=["GET"])
def status():
    """API：返回当前提交状态（供前端轮询或管理页使用）"""
    return jsonify({
        "deadline": DEADLINE.strftime("%Y-%m-%d %H:%M:%S"),
        "deadline_passed": is_deadline_passed(),
        "submission_count": get_submission_count(),
        "server_time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    })


@app.route("/admin", methods=["GET"])
def admin():
    """管理页面：查看所有提交记录"""
    records = get_all_submissions()
    passed = is_deadline_passed()
    return render_template(
        "admin.html",
        records=records,
        deadline_passed=passed,
        deadline_str=DEADLINE.strftime("%Y年%m月%d日 %H:%M"),
        submission_count=len(records),
    )


@app.route("/admin/download", methods=["GET"])
def download_excel():
    """下载提交记录Excel文件"""
    if not os.path.exists(EXCEL_PATH):
        abort(404, description="提交记录文件不存在")
    return send_file(
        EXCEL_PATH,
        as_attachment=True,
        download_name=f"AI作业提交记录_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/uploads/<path:filename>", methods=["GET"])
def uploaded_file(filename):
    """访问上传的附件"""
    return send_file(os.path.join(UPLOAD_DIR, filename))


# ============ 主入口 ============
if __name__ == "__main__":
    print("=" * 60)
    print("  AI作业收集系统")
    print(f"  截止时间：{DEADLINE.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  上传目录：{UPLOAD_DIR}")
    print(f"  数据文件：{EXCEL_PATH}")
    print("=" * 60)
    init_excel()
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
