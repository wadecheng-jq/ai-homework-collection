#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
AI作业提交核对脚本
功能：比对花名册 vs 实际提交记录，生成未提交人员清单
作者：WorkBuddy for Wade
日期：2026-06-05
"""

import os
import sys
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============ 配置区（按需修改）====================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
ROSTER_PATH = os.path.join(BASE_DIR, "..", "..", "..", "Desktop", "花名册.xlsx")
# 如果上面路径不对，请手动修改为花名册的实际路径
ROSTER_SHEET = "Sheet1"  # 花名册工作表名
ROSTER_NAME_COL = 2         # 姓名所在列（1-based，2 = B列）

SUBMISSIONS_PATH = os.path.join(BASE_DIR, "submissions.xlsx")
OUTPUT_DIR = BASE_DIR

# 输出文件名前缀
OUTPUT_PREFIX = "未提交人员清单"

# ================================================


def normalize_name(name):
    """姓名标准化：去空格、全角转半角、统一大小写"""
    if not name:
        return ""
    import unicodedata
    # 全角转半角
    name = unicodedata.normalize("NFKC", str(name).strip())
    name = name.replace(" ", "").replace("　", "")
    return name.lower()


def load_roster(path):
    """加载花名册，返回姓名列表（去重）"""
    if not os.path.exists(path):
        print(f"[ERROR] 花名册文件不存在：{path}")
        print("请修改脚本中的 ROSTER_PATH 为正确路径！")
        sys.exit(1)

    print(f"[INFO] 读取花名册：{path}")
    wb = load_workbook(path, data_only=True)
    ws = wb[ROSTER_SHEET] if ROSTER_SHEET in wb.sheetnames else wb.active

    names = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        if ROSTER_NAME_COL <= len(row):
            name = row[ROSTER_NAME_COL - 1]
            if name:
                names.append(str(name).strip())

    print(f"[INFO] 花名册共 {len(names)} 条记录（含可能的重复）")
    return names


def load_submissions(path):
    """加载已提交记录，返回已提交姓名集合"""
    if not os.path.exists(path):
        print(f"[WARN] 提交记录文件不存在：{path}")
        print("[WARN] 将认为所有人都未提交")
        return set(), []

    print(f"[INFO] 读取提交记录：{path}")
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    submitted_names = set()
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        name = row[2]  # C列 = 姓名
        group = row[3] if len(row) > 3 else ""   # D列 = 圈子/部门
        submit_time = row[1]  # B列 = 提交时间
        if name:
            submitted_names.add(normalize_name(name))
            records.append({
                "seq": row[0],
                "time": submit_time,
                "name": str(name).strip(),
                "group": str(group).strip() if group else "",
            })
        # 也记录序号的姓名（防止同音字问题，保留原始记录）
    print(f"[INFO] 已提交 {len(submitted_names)} 人（去重后）")
    return submitted_names, records


def compare(roster_names, submitted_names):
    """
    比对花名册和提交记录
    返回：(not_submitted_list, duplicate_in_roster)
    """
    # 检查花名册内重复
    seen = set()
    duplicates = []
    unique_roster = []
    for name in roster_names:
        norm = normalize_name(name)
        if not norm:
            continue
        if norm in seen:
            duplicates.append(name)
        else:
            seen.add(norm)
            unique_roster.append(name)

    if duplicates:
        print(f"[WARN] 花名册中发现重复姓名：{', '.join(set(duplicates))}")

    # 比对
    not_submitted = []
    for name in unique_roster:
        if normalize_name(name) not in submitted_names:
            not_submitted.append(name)

    return not_submitted, duplicates, len(unique_roster)


def export_results(not_submitted, roster_count, submitted_count, duplicates, output_path):
    """导出未提交清单到Excel + 生成可转发文本"""
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_path = os.path.join(output_path, f"{OUTPUT_PREFIX}_{timestamp}.xlsx")
    txt_path = os.path.join(output_path, f"{OUTPUT_PREFIX}_{timestamp}.txt")

    # ============ 写Excel ============
    wb = Workbook()
    ws = wb.active
    ws.title = "未提交人员清单"

    # 样式
    header_fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    header_font = Font(name='微软雅黑', bold=True, color="FFFFFF", size=11)
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 汇总信息（顶部说明区）
    ws.merge_cells("A1:D1")
    ws["A1"] = f"AI作业未提交人员清单（生成时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}）"
    ws["A1"].font = Font(name='微软雅黑', bold=True, size=12, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    ws["A2"] = "花名册总人数"
    ws["B2"] = roster_count
    ws["C2"] = "已提交人数"
    ws["D2"] = submitted_count
    ws["A3"] = "未提交人数"
    ws["B3"] = len(not_submitted)
    ws["C3"] = "提交率"
    ws["D3"] = f"{submitted_count/roster_count*100:.1f}%" if roster_count > 0 else "N/A"

    for row in range(2, 4):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = Font(name='微软雅黑', size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20

    # 设置列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 15

    # 表头（第5行起）
    headers = ["序号", "姓名", "备注", "是否已补交"]
    col_widths = [8, 20, 40, 15]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=5, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[5].height = 26

    # 数据行
    for i, name in enumerate(not_submitted, 1):
        row = 5 + i
        ws.cell(row=row, column=1, value=i)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="")
        for col_idx in range(1, 5):
            cell = ws.cell(row=row, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal="center" if col_idx != 3 else "left", vertical="center")
            cell.font = Font(name='微软雅黑', size=10)
        if i % 2 == 0:
            for col_idx in range(1, 5):
                ws.cell(row=row, column=col_idx).fill = PatternFill(
                    start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                )
        ws.row_dimensions[row].height = 22

    wb.save(excel_path)
    print(f"[INFO] 未提交清单已导出：{excel_path}")

    # ============ 写文本文件（可转发）====================
    lines = []
    lines.append("=" * 50)
    lines.append("【AI作业未提交人员通知】")
    lines.append(f"统计时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"花名册人数：{roster_count} | 已提交：{submitted_count} | 未提交：{len(not_submitted)}")
    lines.append(f"提交率：{submitted_count/roster_count*100:.1f}%" if roster_count > 0 else "提交率：N/A")
    lines.append("=" * 50)
    lines.append("")
    if not_submitted:
        lines.append("【未提交名单】")
        for i, name in enumerate(not_submitted, 1):
            lines.append(f"{i}. {name}")
    else:
        lines.append("🎉 所有人已提交，无需催办！")
    lines.append("")
    lines.append("=" * 50)
    lines.append("请各位尽快提交，截止时间：2026-06-08 18:00")
    lines.append("=" * 50)

    txt_content = "\n".join(lines)
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(txt_content)
    print(f"[INFO] 可转发文本已导出：{txt_path}")

    return excel_path, txt_path, txt_content


def main():
    print("=" * 60)
    print("  AI作业提交核对脚本")
    print(f"  生成时间：{datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    # 1. 加载花名册
    roster_names = load_roster(ROSTER_PATH)

    # 2. 加载提交记录
    submitted_names, submitted_records = load_submissions(SUBMISSIONS_PATH)

    # 3. 比对
    not_submitted, duplicates, roster_count = compare(roster_names, submitted_names)

    # 4. 输出结果
    print("\n" + "=" * 60)
    print(f"📊 核对结果")
    print(f"   花名册总人数：{roster_count}")
    print(f"   已提交人数：  {len(submitted_records)}（去重后 {len(submitted_names)} 人）")
    print(f"   未提交人数：  {len(not_submitted)}")
    if roster_count > 0:
        rate = len(submitted_names) / roster_count * 100
        print(f"   提交率：      {rate:.1f}%")
    print("=" * 60)

    if not_submitted:
        print("\n【未提交人员名单】")
        for i, name in enumerate(not_submitted, 1):
            print(f"  {i}. {name}")
    else:
        print("\n🎉 所有人已提交！")

    # 5. 导出文件
    print()
    excel_path, txt_path, txt_content = export_results(
        not_submitted, roster_count, len(submitted_names), duplicates, OUTPUT_DIR
    )

    # 6. 输出可转发文本
    print("\n" + "=" * 60)
    print("【可转发文本】（也可查看导出的 .txt 文件）")
    print("=" * 60)
    print(txt_content)
    print("=" * 60)

    print(f"\n✅ 核对完成！")
    print(f"   Excel清单：{excel_path}")
    print(f"   文本文件：{txt_path}")
    print(f"\n提示：可将 .txt 文件内容直接复制发送到群/私信催办。")


if __name__ == "__main__":
    main()
